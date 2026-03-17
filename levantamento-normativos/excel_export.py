"""Excel/XLSX export via openpyxl for normativo search results.

Generates a formatted workbook with:
- Merged title row showing the search topic
- Green header row with white bold text and auto-filter
- Data rows with per-column formatting (hyperlinks, percentages, wrapping)
- Freeze panes so headers remain visible when scrolling
- Landscape print layout fit to page width

The returned BytesIO buffer is ready for direct use with
``st.download_button(data=buffer)``.
"""

from __future__ import annotations

import logging
import re
from io import BytesIO

from openpyxl import Workbook

logger = logging.getLogger(__name__)
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from models import NormativoResult

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

# -- Camara dos Deputados institutional green --
HEADER_FILL = PatternFill(start_color="4A8C4A", end_color="4A8C4A", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
HEADER_ALIGNMENT = Alignment(
    horizontal="center", vertical="center", wrap_text=True
)

# -- Title row (merged across all columns) --
TITLE_FONT = Font(bold=True, size=14, color="2E7D32", name="Calibri")
TITLE_ALIGNMENT = Alignment(horizontal="center", vertical="center")
TITLE_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

# -- Data row styles --
DATA_FONT = Font(size=10, name="Calibri")
DATA_ALIGNMENT = Alignment(vertical="top", wrap_text=False)
EMENTA_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
LINK_FONT = Font(color="2E7D32", underline="single", size=10, name="Calibri")
NOME_FONT = Font(bold=True, size=10, name="Calibri")
NOME_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
RELEVANCIA_ALIGNMENT = Alignment(horizontal="center", vertical="top")

# -- Relevancia conditional fills --
RELEVANCIA_HIGH_FILL = PatternFill(
    start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"
)  # green, >= 70%
RELEVANCIA_MED_FILL = PatternFill(
    start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"
)  # yellow, >= 40%

# -- Thin border for all cells --
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# -- Column definitions: (header_text, width, source_field) --
COLUMNS = [
    ("Nome do Normativo", 50, "nome"),
    ("Tipo", 20, "tipo"),
    ("Numero", 15, "numero"),
    ("Data", 14, "data"),
    ("Orgao Emissor", 30, "orgao_emissor"),
    ("Ementa", 60, "ementa"),
    ("Link", 40, "link"),
    ("Categoria/Tema", 25, "categoria"),
    ("Situacao", 15, "situacao"),
    ("Relevancia", 12, "relevancia"),
]

# Maximum ementa length in Excel cells. Longer values are truncated to
# prevent workbook bloat and cell rendering issues in older Excel versions.
_MAX_EMENTA_LENGTH = 5000


# ---------------------------------------------------------------------------
# Helper: Date formatting
# ---------------------------------------------------------------------------


def _format_date(date_str: str) -> str:
    """Normalize a date string to DD/MM/YYYY format.

    Handles common input formats:
    - YYYY-MM-DD (ISO format from APIs)
    - DD/MM/YYYY (already correct)
    - YYYY-MM-DDT... (ISO datetime, truncated to date)
    - Empty/None -> empty string

    Args:
        date_str: Raw date string from NormativoResult.data.

    Returns:
        Date formatted as DD/MM/YYYY, or the original string if parsing fails,
        or empty string if input is empty.
    """
    if not date_str:
        return ""

    date_str = date_str.strip()

    # Already in DD/MM/YYYY format
    if re.match(r"^\d{2}/\d{2}/\d{4}$", date_str):
        return date_str

    # ISO format: YYYY-MM-DD or YYYY-MM-DDT...
    iso_match = re.match(r"^(\d{4})-(\d{2})-(\d{2})", date_str)
    if iso_match:
        year, month, day = iso_match.groups()
        return f"{day}/{month}/{year}"

    # Unrecognized format: return as-is rather than losing data
    return date_str


# ---------------------------------------------------------------------------
# Helper: Title row
# ---------------------------------------------------------------------------


def _write_title_row(ws, topic: str, num_columns: int) -> None:
    """Write the merged title row at row 1.

    Args:
        ws: Active worksheet.
        topic: Search topic to display in the title.
        num_columns: Total number of columns to merge across.
    """
    last_col = get_column_letter(num_columns)
    ws.merge_cells(f"A1:{last_col}1")

    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"Levantamento de Normativos: {topic}"
    title_cell.font = TITLE_FONT
    title_cell.alignment = TITLE_ALIGNMENT
    title_cell.fill = TITLE_FILL

    ws.row_dimensions[1].height = 40


# ---------------------------------------------------------------------------
# Helper: Header row
# ---------------------------------------------------------------------------


def _write_header_row(ws) -> None:
    """Write the formatted header row at row 2.

    Args:
        ws: Active worksheet.
    """
    for col_idx, (header_text, _, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header_text
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    ws.row_dimensions[2].height = 30


# ---------------------------------------------------------------------------
# Helper: Data row
# ---------------------------------------------------------------------------


def _write_data_row(ws, row_idx: int, item: NormativoResult) -> None:
    """Write a single data row for one NormativoResult.

    Handles per-column formatting including:
    - Date formatting (DD/MM/YYYY)
    - Ementa truncation and text wrapping
    - Hyperlink creation for the Link column
    - Percentage formatting and conditional fill for Relevancia

    Args:
        ws: Active worksheet.
        row_idx: The 1-based row number to write to.
        item: The NormativoResult to render.
    """
    for col_idx, (_, _, field_name) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=row_idx, column=col_idx)
        raw_value = getattr(item, field_name, "")
        # Preserve numeric 0.0 for relevancia; the generic `or ""` guard
        # would convert falsy 0.0 to "", breaking the round-trip.
        value = raw_value if field_name == "relevancia" else (raw_value or "")

        # -- Per-field formatting --

        if field_name == "data":
            # Format date as DD/MM/YYYY string. The data field is already
            # a string in the model; we just ensure consistent display.
            cell.value = _format_date(value)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT

        elif field_name == "ementa":
            # Truncate very long ementas to prevent workbook bloat
            if len(value) > _MAX_EMENTA_LENGTH:
                logger.info(
                    "Ementa truncated from %d to %d chars for '%s'.",
                    len(value), _MAX_EMENTA_LENGTH,
                    getattr(item, "nome", "unknown")[:80],
                )
                value = value[:_MAX_EMENTA_LENGTH] + "..."
            cell.value = value
            cell.font = DATA_FONT
            cell.alignment = EMENTA_ALIGNMENT
            # Set row height for rows with substantial ementas
            if len(value) > 100:
                ws.row_dimensions[row_idx].height = 60

        elif field_name == "link":
            # Create clickable hyperlink
            cell.value = value
            if value and value.startswith(("http://", "https://")):
                cell.hyperlink = value
            cell.font = LINK_FONT
            cell.alignment = DATA_ALIGNMENT

        elif field_name == "relevancia":
            # Format as percentage with conditional coloring.
            # value may be a float (including 0.0) or None/empty string.
            numeric_value = float(value) if value is not None and value != "" else 0.0
            cell.value = numeric_value
            cell.number_format = "0%"
            cell.font = DATA_FONT
            cell.alignment = RELEVANCIA_ALIGNMENT

            # Conditional fill based on relevance score
            if numeric_value >= 0.7:
                cell.fill = RELEVANCIA_HIGH_FILL
            elif numeric_value >= 0.4:
                cell.fill = RELEVANCIA_MED_FILL

        elif field_name == "nome":
            cell.value = value
            cell.font = NOME_FONT
            cell.alignment = NOME_ALIGNMENT

        else:
            # Default: plain text cell
            cell.value = value
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT

        cell.border = THIN_BORDER


# ---------------------------------------------------------------------------
# Main export function
# ---------------------------------------------------------------------------


def generate_excel(results: list[NormativoResult], topic: str) -> BytesIO:
    """Generate a formatted Excel workbook from selected normativos.

    The workbook contains a single sheet named 'Normativos' with:
    - Row 1: merged title row showing the search topic
    - Row 2: green header row with white bold text
    - Rows 3+: data rows with formatting per column type
    - Freeze panes on row 3 (headers always visible)
    - Auto-filter on the header row
    - Hyperlinks in the Link column
    - Conditional formatting on the Relevancia column

    If the results list is empty, the workbook is returned with only the
    title and header rows (no data rows).

    Args:
        results: List of NormativoResult objects to export. May be empty,
                 in which case a workbook with just headers is returned.
        topic: The search topic string, displayed in the title row and used
               to contextualize the workbook for stakeholders.

    Returns:
        BytesIO buffer containing the complete .xlsx file. The buffer is
        seeked to position 0, ready for direct use with
        st.download_button(data=buffer).
    """
    logger.info("Generating Excel with %d results for topic '%s'.", len(results), topic[:80])

    wb = Workbook()
    ws = wb.active
    ws.title = "Normativos"

    num_columns = len(COLUMNS)

    # ----------------------------------------------------------------
    # Row 1: Title row (merged across all columns)
    # ----------------------------------------------------------------
    _write_title_row(ws, topic, num_columns)

    # ----------------------------------------------------------------
    # Row 2: Header row
    # ----------------------------------------------------------------
    _write_header_row(ws)

    # ----------------------------------------------------------------
    # Rows 3+: Data rows
    # ----------------------------------------------------------------
    for row_idx, item in enumerate(results, start=3):
        _write_data_row(ws, row_idx, item)

    # ----------------------------------------------------------------
    # Column widths
    # ----------------------------------------------------------------
    for col_idx, (_, width, _) in enumerate(COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    # ----------------------------------------------------------------
    # Freeze panes: freeze below row 2 (title + header always visible)
    # ----------------------------------------------------------------
    ws.freeze_panes = "A3"

    # ----------------------------------------------------------------
    # Auto-filter on header row (row 2)
    # ----------------------------------------------------------------
    last_col_letter = get_column_letter(num_columns)
    last_data_row = len(results) + 2  # +2 for title and header rows
    ws.auto_filter.ref = f"A2:{last_col_letter}{last_data_row}"

    # ----------------------------------------------------------------
    # Print settings for physical printing
    # ----------------------------------------------------------------
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # 0 = as many pages as needed vertically

    # ----------------------------------------------------------------
    # Write to BytesIO buffer
    # ----------------------------------------------------------------
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
