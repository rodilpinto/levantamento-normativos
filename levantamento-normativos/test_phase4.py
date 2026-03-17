"""Comprehensive tests for Phase 4: Deduplication and Excel Export.

Tests the deduplicator module (cross-source deduplication with three strategies)
and the excel_export module (formatted .xlsx generation via openpyxl).

Run from within levantamento-normativos/:
    python -m pytest test_phase4.py -v
"""

from __future__ import annotations

import hashlib
from io import BytesIO

import pytest
from openpyxl import load_workbook

from deduplicator import _normalize, _merge, deduplicate
from excel_export import generate_excel, COLUMNS, _MAX_EMENTA_LENGTH
from models import NormativoResult


# ---------------------------------------------------------------------------
# Helpers: factory functions for building test NormativoResult instances
# ---------------------------------------------------------------------------


def _make_result(
    nome: str = "Lei n. 13.709/2018",
    tipo: str = "Lei",
    numero: str = "13.709",
    data: str | None = "14/08/2018",
    orgao_emissor: str = "Presidencia da Republica",
    ementa: str = "Dispoe sobre a protecao de dados pessoais.",
    link: str = "https://www.planalto.gov.br/lei13709",
    source: str = "lexml",
    found_by: str = "LGPD",
    categoria: str = "Protecao de Dados",
    situacao: str = "Vigente",
    relevancia: float = 0.8,
) -> NormativoResult:
    """Shorthand factory that creates a NormativoResult with sensible defaults."""
    return NormativoResult(
        nome=nome,
        tipo=tipo,
        numero=numero,
        data=data,
        orgao_emissor=orgao_emissor,
        ementa=ementa,
        link=link,
        source=source,
        found_by=found_by,
        categoria=categoria,
        situacao=situacao,
        relevancia=relevancia,
    )


# ===========================================================================
#  DEDUPLICATOR TESTS
# ===========================================================================


class TestDeduplicatorExactIdMatch:
    """Strategy 1: Exact ID (SHA-256 of tipo|numero|data) match."""

    def test_same_tipo_numero_data_deduplicates_to_one(self):
        """Two results with identical tipo+numero+data produce the same ID
        and should be merged into a single output record."""
        r1 = _make_result(source="lexml", found_by="LGPD")
        r2 = _make_result(source="google", found_by="protecao dados")

        # Sanity: both have the same auto-generated ID
        assert r1.id == r2.id, "Precondition: IDs must match for this test"

        result = deduplicate([r1, r2])
        assert len(result) == 1, (
            f"Expected 1 result after dedup of exact ID match, got {len(result)}"
        )


class TestDeduplicatorTipoNumeroMatch:
    """Strategy 2: Tipo + Numero case-insensitive match."""

    def test_same_tipo_numero_different_data_merges(self):
        """Same tipo+numero but slightly different date should merge via
        Strategy 2 even though IDs differ."""
        r1 = _make_result(
            tipo="Instrucao Normativa", numero="65", data="10/01/2020",
            source="lexml", found_by="seguranca",
        )
        r2 = _make_result(
            tipo="Instrucao Normativa", numero="65", data="2020-01-10",
            source="tcu", found_by="auditoria",
        )

        # IDs differ because the raw data strings differ
        assert r1.id != r2.id, "Precondition: IDs must differ"

        result = deduplicate([r1, r2])
        assert len(result) == 1, (
            f"Expected 1 result after tipo+numero match, got {len(result)}"
        )

    def test_case_insensitive_tipo_match(self):
        """Tipo matching should be case-insensitive."""
        r1 = _make_result(tipo="LEI", numero="9999", data="01/01/2000", source="lexml")
        r2 = _make_result(tipo="lei", numero="9999", data="02/01/2000", source="tcu")

        assert r1.id != r2.id
        result = deduplicate([r1, r2])
        assert len(result) == 1


class TestDeduplicatorFuzzyEmentaMatch:
    """Strategy 3: Fuzzy ementa comparison via SequenceMatcher."""

    def test_very_similar_ementas_merge(self):
        """Two results with >85% similar ementas but different tipo/numero
        should be merged by fuzzy matching."""
        ementa_v1 = (
            "Dispoe sobre o tratamento de dados pessoais, inclusive nos "
            "meios digitais, por pessoa natural ou por pessoa juridica de "
            "direito publico ou privado, com o objetivo de proteger os "
            "direitos fundamentais de liberdade e de privacidade e o livre "
            "desenvolvimento da personalidade da pessoa natural."
        )
        # Very similar content with minor wording difference, yielding ratio > 0.85
        ementa_v2 = (
            "Dispoe sobre o tratamento de dados pessoais, inclusive nos "
            "meios digitais, por pessoa natural ou por pessoa juridica de "
            "direito publico ou privado, com o objetivo de proteger os "
            "direitos fundamentais de liberdade e de privacidade e o livre "
            "desenvolvimento da personalidade de pessoa natural."
        )

        r1 = _make_result(
            tipo="Lei", numero="13.709", ementa=ementa_v1,
            source="lexml", found_by="dados pessoais",
        )
        r2 = _make_result(
            tipo="Framework", numero="LGPD", ementa=ementa_v2,
            source="google", found_by="LGPD",
        )

        # Ensure they won't match on ID or tipo+numero
        assert r1.id != r2.id
        assert (r1.tipo.lower(), r1.numero.lower()) != (r2.tipo.lower(), r2.numero.lower())

        result = deduplicate([r1, r2])
        assert len(result) == 1, (
            f"Expected fuzzy match to merge similar ementas, got {len(result)} results"
        )


class TestDeduplicatorNoFalsePositives:
    """Genuinely different normativos must remain separate."""

    def test_different_normativos_stay_separate(self):
        """Two completely different normativos should not be merged."""
        r1 = _make_result(
            nome="Lei Geral de Protecao de Dados",
            tipo="Lei", numero="13.709", data="14/08/2018",
            ementa="Dispoe sobre a protecao de dados pessoais.",
            source="lexml",
        )
        r2 = _make_result(
            nome="Marco Civil da Internet",
            tipo="Lei", numero="12.965", data="23/04/2014",
            ementa="Estabelece principios, garantias, direitos e deveres "
                   "para o uso da Internet no Brasil.",
            source="lexml",
        )

        result = deduplicate([r1, r2])
        assert len(result) == 2, (
            f"Expected 2 distinct results, got {len(result)}"
        )


class TestDeduplicatorMergeLogic:
    """Verify that merged records combine the best metadata from both sources."""

    def test_source_is_combined(self):
        r1 = _make_result(source="lexml", found_by="LGPD")
        r2 = _make_result(source="google", found_by="LGPD")

        result = deduplicate([r1, r2])
        sources = set(s.strip() for s in result[0].source.split(","))
        assert "lexml" in sources
        assert "google" in sources

    def test_found_by_is_combined(self):
        r1 = _make_result(source="lexml", found_by="LGPD")
        r2 = _make_result(source="google", found_by="protecao dados")

        result = deduplicate([r1, r2])
        keywords = set(k.strip() for k in result[0].found_by.split(","))
        assert "LGPD" in keywords
        assert "protecao dados" in keywords

    def test_longer_ementa_is_kept(self):
        short_ementa = "Dispoe sobre dados pessoais."
        long_ementa = (
            "Dispoe sobre a protecao de dados pessoais, inclusive nos "
            "meios digitais, por pessoa natural ou por pessoa juridica."
        )
        r1 = _make_result(ementa=short_ementa, source="lexml")
        r2 = _make_result(ementa=long_ementa, source="google")

        result = deduplicate([r1, r2])
        assert result[0].ementa == long_ementa

    def test_higher_relevancia_is_kept(self):
        r1 = _make_result(relevancia=0.5, source="lexml")
        r2 = _make_result(relevancia=0.9, source="google")

        result = deduplicate([r1, r2])
        assert result[0].relevancia == 0.9

    def test_link_from_authoritative_source_preferred(self):
        """lexml link should be preferred over google link."""
        r1 = _make_result(
            source="google",
            link="https://google.com/result",
        )
        r2 = _make_result(
            source="lexml",
            link="https://lexml.gov.br/lei13709",
        )

        result = deduplicate([r1, r2])
        # After merge, source is combined (sorted alphabetically: "google, lexml").
        # The _merge function checks source priority AFTER updating source,
        # so the merged source string starts with "google" alphabetically.
        # The incoming (lexml, priority 0) has lower priority number than
        # the existing's first source in the combined string.
        # Let's just verify the link: lexml should win.
        # NOTE: There is a subtlety in the implementation -- after merge,
        # existing.source becomes "google, lexml" so existing_priority looks
        # up "google" (priority 2), incoming_priority looks up the incoming
        # source. But incoming.source is "lexml" so incoming_priority = 0.
        # Since 0 < 2, the incoming link wins. Good.
        assert result[0].link == "https://lexml.gov.br/lei13709"


class TestDeduplicatorEdgeCases:
    """Edge cases: empty list, single item, three-way merge."""

    def test_empty_list(self):
        assert deduplicate([]) == []

    def test_single_item(self):
        r1 = _make_result()
        result = deduplicate([r1])
        assert len(result) == 1
        assert result[0] is r1

    def test_three_way_merge(self):
        """Same normativo from all 3 sources should merge to 1 result."""
        r1 = _make_result(source="lexml", found_by="LGPD", relevancia=0.7)
        r2 = _make_result(source="tcu", found_by="dados pessoais", relevancia=0.8)
        r3 = _make_result(source="google", found_by="lei dados", relevancia=0.6)

        result = deduplicate([r1, r2, r3])
        assert len(result) == 1

        merged = result[0]
        sources = set(s.strip() for s in merged.source.split(","))
        assert sources == {"lexml", "tcu", "google"}

        keywords = set(k.strip() for k in merged.found_by.split(","))
        assert "LGPD" in keywords
        assert "dados pessoais" in keywords
        assert "lei dados" in keywords

        assert merged.relevancia == 0.8  # max of 0.7, 0.8, 0.6


class TestNormalizeHelper:
    """Tests for the _normalize() text normalization function."""

    def test_strips_accents(self):
        assert "regulamentacao" in _normalize("Regulamentação")

    def test_lowercases(self):
        assert _normalize("ABC DEF") == "abc def"

    def test_collapses_whitespace(self):
        assert _normalize("a   b\t\nc") == "a b c"

    def test_removes_punctuation(self):
        assert _normalize("lei (n. 13.709)") == "lei n 13709"

    def test_empty_string(self):
        assert _normalize("") == ""

    def test_none_like(self):
        # _normalize expects a string; empty string is the None guard
        assert _normalize("") == ""


# ===========================================================================
#  EXCEL EXPORT TESTS
# ===========================================================================


def _load_workbook_from_buffer(buf: BytesIO):
    """Helper: load an openpyxl Workbook from a generate_excel buffer."""
    return load_workbook(BytesIO(buf.getvalue()))


def _make_sample_results(count: int = 3) -> list[NormativoResult]:
    """Create a list of distinct NormativoResult objects for Excel tests."""
    samples = [
        _make_result(
            nome="Lei n. 13.709/2018 - LGPD",
            tipo="Lei", numero="13.709", data="14/08/2018",
            ementa="Dispoe sobre a protecao de dados pessoais.",
            link="https://www.planalto.gov.br/lei13709",
            source="lexml", relevancia=0.85,
        ),
        _make_result(
            nome="Decreto n. 10.332/2020",
            tipo="Decreto", numero="10.332", data="28/04/2020",
            ementa="Institui a Estrategia de Governo Digital.",
            link="https://www.planalto.gov.br/decreto10332",
            source="tcu", relevancia=0.6,
        ),
        _make_result(
            nome="IN SGD/ME n. 1/2019",
            tipo="Instrucao Normativa", numero="1", data="04/04/2019",
            ementa="Dispoe sobre o processo de contratacao de TIC.",
            link="https://www.gov.br/in1-2019",
            source="google", relevancia=0.3,
        ),
    ]
    return samples[:count]


class TestExcelBasicExport:
    """Basic Excel generation and format verification."""

    def test_basic_export_produces_valid_xlsx(self):
        """Generated buffer should start with PK (zip/xlsx magic bytes)."""
        results = _make_sample_results(3)
        buf = generate_excel(results, "Governanca de TI")

        raw = buf.getvalue()
        assert raw[:2] == b"PK", (
            f"Expected xlsx magic bytes PK, got {raw[:2]!r}"
        )

    def test_basic_export_has_correct_row_count(self):
        """3 results -> row 1 (title) + row 2 (header) + 3 data rows = 5 rows."""
        results = _make_sample_results(3)
        buf = generate_excel(results, "test")
        wb = _load_workbook_from_buffer(buf)
        ws = wb.active

        # max_row should be 5 (title + header + 3 data)
        assert ws.max_row == 5, f"Expected 5 rows, got {ws.max_row}"

    def test_title_row_contains_topic(self):
        results = _make_sample_results(1)
        buf = generate_excel(results, "Seguranca da Informacao")
        wb = _load_workbook_from_buffer(buf)
        ws = wb.active

        title = ws.cell(row=1, column=1).value
        assert "Seguranca da Informacao" in title

    def test_sheet_name(self):
        results = _make_sample_results(1)
        buf = generate_excel(results, "test")
        wb = _load_workbook_from_buffer(buf)
        assert wb.active.title == "Normativos"


class TestExcelEmptyList:
    """generate_excel with empty results list."""

    def test_empty_list_returns_valid_xlsx(self):
        """The actual implementation returns a valid workbook with headers only
        (no ValueError is raised despite the spec suggesting otherwise)."""
        buf = generate_excel([], "test")
        raw = buf.getvalue()
        assert raw[:2] == b"PK", "Empty list should still produce valid xlsx"

    def test_empty_list_has_header_row(self):
        buf = generate_excel([], "test")
        wb = _load_workbook_from_buffer(buf)
        ws = wb.active

        # Row 2 should have header values
        header_val = ws.cell(row=2, column=1).value
        assert header_val == "Nome do Normativo"


class TestExcelHyperlinks:
    """Verify the link column contains hyperlink formatting."""

    def test_link_column_has_hyperlink(self):
        results = _make_sample_results(1)
        buf = generate_excel(results, "test")
        wb = _load_workbook_from_buffer(buf)
        ws = wb.active

        # Link is column 7 (index in COLUMNS)
        link_cell = ws.cell(row=3, column=7)
        assert link_cell.value == results[0].link
        assert link_cell.hyperlink is not None, "Link cell should have a hyperlink"
        # Hyperlink target should match the URL
        hyperlink_target = link_cell.hyperlink.target
        assert hyperlink_target == results[0].link, (
            f"Hyperlink target {hyperlink_target!r} != expected {results[0].link!r}"
        )


class TestExcelColumnCount:
    """Verify all 10 expected columns are present."""

    def test_has_10_columns(self):
        assert len(COLUMNS) == 10, f"Expected 10 column definitions, got {len(COLUMNS)}"

    def test_header_row_has_10_columns(self):
        results = _make_sample_results(1)
        buf = generate_excel(results, "test")
        wb = _load_workbook_from_buffer(buf)
        ws = wb.active

        headers = []
        for col in range(1, 11):
            val = ws.cell(row=2, column=col).value
            if val:
                headers.append(val)

        assert len(headers) == 10, f"Expected 10 headers, got {len(headers)}: {headers}"

    def test_expected_header_names(self):
        expected = [
            "Nome do Normativo", "Tipo", "Numero", "Data", "Orgao Emissor",
            "Ementa", "Link", "Categoria/Tema", "Situacao", "Relevancia",
        ]
        results = _make_sample_results(1)
        buf = generate_excel(results, "test")
        wb = _load_workbook_from_buffer(buf)
        ws = wb.active

        actual = [ws.cell(row=2, column=c).value for c in range(1, 11)]
        assert actual == expected, f"Header mismatch: {actual}"


class TestExcelLargeEmenta:
    """NormativoResult with a very long ementa should be truncated."""

    def test_6000_char_ementa_is_truncated(self):
        long_ementa = "A" * 6000
        r = _make_result(ementa=long_ementa)
        buf = generate_excel([r], "test")
        wb = _load_workbook_from_buffer(buf)
        ws = wb.active

        ementa_cell = ws.cell(row=3, column=6)  # Ementa is column 6
        cell_value = ementa_cell.value

        # Should be truncated to _MAX_EMENTA_LENGTH + "..."
        assert len(cell_value) == _MAX_EMENTA_LENGTH + 3, (
            f"Expected truncated length {_MAX_EMENTA_LENGTH + 3}, got {len(cell_value)}"
        )
        assert cell_value.endswith("..."), "Truncated ementa should end with '...'"

    def test_short_ementa_is_not_truncated(self):
        short = "Dispoe sobre dados pessoais."
        r = _make_result(ementa=short)
        buf = generate_excel([r], "test")
        wb = _load_workbook_from_buffer(buf)
        ws = wb.active

        assert ws.cell(row=3, column=6).value == short


class TestExcelSpecialCharacters:
    """Ementas with accents, quotes, and newlines should not crash export."""

    def test_accented_characters(self):
        ementa = "Regulamentação sobre proteção de dados — incluindo ações específicas."
        r = _make_result(ementa=ementa)
        buf = generate_excel([r], "test")
        wb = _load_workbook_from_buffer(buf)
        ws = wb.active
        assert ws.cell(row=3, column=6).value == ementa

    def test_quotes_and_special_chars(self):
        ementa = 'Dispõe sobre "normas" de \'segurança\' & governança <TI>.'
        r = _make_result(ementa=ementa)
        # Should not raise
        buf = generate_excel([r], "test")
        wb = _load_workbook_from_buffer(buf)
        ws = wb.active
        assert ws.cell(row=3, column=6).value == ementa

    def test_newlines_in_ementa(self):
        ementa = "Linha 1.\nLinha 2.\nLinha 3."
        r = _make_result(ementa=ementa)
        buf = generate_excel([r], "test")
        wb = _load_workbook_from_buffer(buf)
        ws = wb.active
        assert ws.cell(row=3, column=6).value == ementa

    def test_mixed_unicode(self):
        ementa = "§ 1° — Art. 5°, inciso XII, da Constituição Federal de 1988."
        r = _make_result(ementa=ementa)
        buf = generate_excel([r], "test")
        wb = _load_workbook_from_buffer(buf)
        ws = wb.active
        assert ws.cell(row=3, column=6).value == ementa


class TestExcelDataValues:
    """Round-trip verification: values written match values read back."""

    def test_nome_value(self):
        r = _make_result(nome="Lei n. 13.709/2018 - LGPD")
        buf = generate_excel([r], "test")
        wb = _load_workbook_from_buffer(buf)
        ws = wb.active
        assert ws.cell(row=3, column=1).value == "Lei n. 13.709/2018 - LGPD"

    def test_tipo_value(self):
        r = _make_result(tipo="Instrucao Normativa")
        buf = generate_excel([r], "test")
        wb = _load_workbook_from_buffer(buf)
        ws = wb.active
        assert ws.cell(row=3, column=2).value == "Instrucao Normativa"

    def test_relevancia_is_numeric(self):
        r = _make_result(relevancia=0.85)
        buf = generate_excel([r], "test")
        wb = _load_workbook_from_buffer(buf)
        ws = wb.active
        val = ws.cell(row=3, column=10).value
        assert isinstance(val, float), f"Relevancia should be float, got {type(val)}"
        assert abs(val - 0.85) < 0.001

    def test_date_formatting_iso_to_ddmmyyyy(self):
        """ISO date input (2020-04-28) should be formatted as 28/04/2020."""
        r = _make_result(data="2020-04-28")
        buf = generate_excel([r], "test")
        wb = _load_workbook_from_buffer(buf)
        ws = wb.active
        assert ws.cell(row=3, column=4).value == "28/04/2020"

    def test_date_already_ddmmyyyy(self):
        r = _make_result(data="14/08/2018")
        buf = generate_excel([r], "test")
        wb = _load_workbook_from_buffer(buf)
        ws = wb.active
        assert ws.cell(row=3, column=4).value == "14/08/2018"


class TestExcelFreezePanes:
    """Verify freeze panes are set correctly."""

    def test_freeze_at_a3(self):
        results = _make_sample_results(1)
        buf = generate_excel(results, "test")
        wb = _load_workbook_from_buffer(buf)
        ws = wb.active
        assert ws.freeze_panes == "A3", (
            f"Expected freeze_panes='A3', got {ws.freeze_panes!r}"
        )


# ===========================================================================
#  Run via pytest or direct execution
# ===========================================================================

if __name__ == "__main__":
    pytest.main([__file__, "-v"])
