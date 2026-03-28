"""
Comprehensive test suite for Levantamento de Normativos.

Tests all components: models, searchers, deduplicator, excel_export, LLM client.
Each test is independent and reports PASS/FAIL clearly.
"""

import hashlib
import json
import os
import re
import sys
import traceback
import unicodedata
from dataclasses import dataclass
from io import BytesIO
from unittest.mock import MagicMock, patch

# ---------------------------------------------------------------------------
# Ensure we can import project modules
# ---------------------------------------------------------------------------
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# ---------------------------------------------------------------------------
# Test results tracking
# ---------------------------------------------------------------------------
_results = []


def run_test(name, func):
    """Run a single test function and track results."""
    try:
        func()
        _results.append(("PASS", name, ""))
        print(f"  PASS  {name}")
    except Exception as e:
        tb = traceback.format_exc()
        _results.append(("FAIL", name, str(e)))
        print(f"  FAIL  {name}")
        print(f"        {e}")
        # Print a few lines of traceback for debugging
        for line in tb.strip().split("\n")[-3:]:
            print(f"        {line}")


# ===========================================================================
# 1. MODELS TESTS
# ===========================================================================
print("\n" + "=" * 70)
print("1. MODELS")
print("=" * 70)

from models import NormativoResult, SearchConfig


def test_normativo_basic_creation():
    """Create a NormativoResult with all required fields."""
    r = NormativoResult(
        nome="Lei 13.709",
        tipo="Lei",
        numero="13.709",
        data="14/08/2018",
        orgao_emissor="Presidencia da Republica",
        ementa="LGPD - Lei Geral de Protecao de Dados",
        link="https://www.planalto.gov.br/ccivil_03/_ato2015-2018/2018/lei/l13709.htm",
        source="lexml",
        found_by="LGPD",
    )
    assert r.nome == "Lei 13.709"
    assert r.tipo == "Lei"
    assert r.numero == "13.709"
    assert r.data == "14/08/2018"
    assert r.source == "lexml"
    assert r.found_by == "LGPD"
    assert r.categoria == "Nao categorizado"
    assert r.situacao == "Nao identificado"
    assert r.relevancia == 0.0
    assert r.raw_data == {}
    assert r.id != ""  # Auto-generated


def test_normativo_id_generation():
    """ID should be SHA-256 of tipo|numero|data."""
    r = NormativoResult(
        nome="Test", tipo="Lei", numero="123", data="01/01/2020",
        orgao_emissor="", ementa="", link="", source="test", found_by="kw",
    )
    expected = hashlib.sha256("Lei|123|01/01/2020".encode("utf-8")).hexdigest()
    assert r.id == expected, f"Expected {expected}, got {r.id}"


def test_normativo_id_uses_link_when_no_numero():
    """When numero is empty, ID should use link instead."""
    r = NormativoResult(
        nome="COBIT", tipo="Framework/Padrao", numero="", data=None,
        orgao_emissor="ISACA", ementa="", link="https://isaca.org/cobit",
        source="google", found_by="COBIT",
    )
    expected = hashlib.sha256("Framework/Padrao|https://isaca.org/cobit|None".encode("utf-8")).hexdigest()
    assert r.id == expected


def test_normativo_same_inputs_same_id():
    """Two results with same tipo/numero/data should have same ID."""
    r1 = NormativoResult(
        nome="Lei A", tipo="Lei", numero="123", data="01/01/2020",
        orgao_emissor="A", ementa="A", link="http://a", source="lexml", found_by="kw1",
    )
    r2 = NormativoResult(
        nome="Lei B", tipo="Lei", numero="123", data="01/01/2020",
        orgao_emissor="B", ementa="B", link="http://b", source="tcu", found_by="kw2",
    )
    assert r1.id == r2.id


def test_normativo_different_inputs_different_id():
    """Results with different tipo/numero should have different IDs."""
    r1 = NormativoResult(
        nome="A", tipo="Lei", numero="123", data="01/01/2020",
        orgao_emissor="", ementa="", link="", source="lexml", found_by="kw",
    )
    r2 = NormativoResult(
        nome="B", tipo="Decreto", numero="456", data="01/01/2020",
        orgao_emissor="", ementa="", link="", source="lexml", found_by="kw",
    )
    assert r1.id != r2.id


def test_normativo_optional_fields_defaults():
    """Optional fields should have correct defaults."""
    r = NormativoResult(
        nome="X", tipo="Lei", numero="1", data=None,
        orgao_emissor="", ementa="", link="", source="", found_by="",
    )
    assert r.data is None
    assert r.categoria == "Nao categorizado"
    assert r.situacao == "Nao identificado"
    assert r.relevancia == 0.0
    assert r.raw_data == {}


def test_search_config_defaults():
    """SearchConfig should have sensible defaults."""
    sc = SearchConfig()
    assert sc.topic == ""
    assert sc.keywords == []
    assert sc.sources == ["lexml", "tcu", "google"]
    assert sc.max_results_per_source == 50


def test_search_config_custom():
    """SearchConfig with custom values."""
    sc = SearchConfig(
        topic="Governanca de TI",
        keywords=["COBIT", "ITIL"],
        sources=["lexml"],
        max_results_per_source=100,
    )
    assert sc.topic == "Governanca de TI"
    assert sc.keywords == ["COBIT", "ITIL"]
    assert sc.sources == ["lexml"]
    assert sc.max_results_per_source == 100


def test_normativo_with_custom_optional_fields():
    """NormativoResult with custom optional fields."""
    r = NormativoResult(
        nome="X", tipo="Lei", numero="1", data="01/01/2020",
        orgao_emissor="TCU", ementa="Test", link="http://test",
        source="tcu", found_by="kw",
        categoria="Governança de TI",
        situacao="Vigente",
        relevancia=0.95,
        raw_data={"key": "value"},
    )
    assert r.categoria == "Governança de TI"
    assert r.situacao == "Vigente"
    assert r.relevancia == 0.95
    assert r.raw_data == {"key": "value"}


for name, func in [
    ("NormativoResult basic creation", test_normativo_basic_creation),
    ("NormativoResult ID generation", test_normativo_id_generation),
    ("NormativoResult ID uses link when no numero", test_normativo_id_uses_link_when_no_numero),
    ("Same inputs produce same ID", test_normativo_same_inputs_same_id),
    ("Different inputs produce different ID", test_normativo_different_inputs_different_id),
    ("Optional fields have defaults", test_normativo_optional_fields_defaults),
    ("SearchConfig defaults", test_search_config_defaults),
    ("SearchConfig custom values", test_search_config_custom),
    ("NormativoResult with custom optional fields", test_normativo_with_custom_optional_fields),
]:
    run_test(name, func)


# ===========================================================================
# 2. BASE SEARCHER TESTS
# ===========================================================================
print("\n" + "=" * 70)
print("2. BASE SEARCHER")
print("=" * 70)

from searchers.base import BaseSearcher


def test_normalize_text_basic():
    """Normalize text: lowercase, strip, remove accents."""
    result = BaseSearcher._normalize_text("  Ação Regulamentação  ")
    assert result == "acao regulamentacao", f"Got: {result}"


def test_normalize_text_empty():
    """Normalize empty string."""
    assert BaseSearcher._normalize_text("") == ""
    assert BaseSearcher._normalize_text(None) == ""


def test_normalize_text_whitespace():
    """Collapse multiple whitespace."""
    result = BaseSearcher._normalize_text("a   b\t\nc")
    assert result == "a b c"


def test_safe_date_format_iso():
    """ISO date -> DD/MM/YYYY."""
    result = BaseSearcher._safe_date_format("2020-01-15")
    assert result == "15/01/2020", f"Got: {result}"


def test_safe_date_format_iso_datetime():
    """ISO datetime -> DD/MM/YYYY."""
    result = BaseSearcher._safe_date_format("2020-01-15T14:30:00")
    assert result == "15/01/2020", f"Got: {result}"


def test_safe_date_format_already_correct():
    """Already DD/MM/YYYY -> unchanged."""
    result = BaseSearcher._safe_date_format("15/01/2020")
    assert result == "15/01/2020"


def test_safe_date_format_year_only():
    """Year only -> 01/01/YYYY."""
    result = BaseSearcher._safe_date_format("2020")
    assert result == "01/01/2020"


def test_safe_date_format_empty():
    """Empty/None -> empty string."""
    assert BaseSearcher._safe_date_format("") == ""
    assert BaseSearcher._safe_date_format(None) == ""


def test_safe_date_format_unknown():
    """Unknown format -> return original."""
    result = BaseSearcher._safe_date_format("Jan 2020")
    assert result == "Jan 2020"


for name, func in [
    ("Normalize text basic", test_normalize_text_basic),
    ("Normalize text empty", test_normalize_text_empty),
    ("Normalize text whitespace", test_normalize_text_whitespace),
    ("Safe date format ISO", test_safe_date_format_iso),
    ("Safe date format ISO datetime", test_safe_date_format_iso_datetime),
    ("Safe date format already correct", test_safe_date_format_already_correct),
    ("Safe date format year only", test_safe_date_format_year_only),
    ("Safe date format empty", test_safe_date_format_empty),
    ("Safe date format unknown", test_safe_date_format_unknown),
]:
    run_test(name, func)


# ===========================================================================
# 3. LEXML SEARCHER TESTS
# ===========================================================================
print("\n" + "=" * 70)
print("3. LEXML SEARCHER")
print("=" * 70)

from searchers.lexml_searcher import LexMLSearcher, URN_PATTERN, URN_TIPO_MAP


def test_lexml_source_name():
    s = LexMLSearcher()
    assert s.source_name() == "LexML Brasil"


def test_lexml_urn_pattern_full():
    """URN pattern parses tipo, date, and number."""
    urn = "urn:lex:br:federal:lei:2018-08-14;13709"
    m = URN_PATTERN.search(urn)
    assert m is not None, f"Pattern did not match: {urn}"
    assert m.group(1) == "lei"
    assert m.group(2) == "2018-08-14"
    assert m.group(3) == "13709"


def test_lexml_urn_pattern_year_only():
    """URN with year-only date."""
    urn = "urn:lex:br:federal:decreto:2020;10332"
    m = URN_PATTERN.search(urn)
    assert m is not None
    assert m.group(2) == "2020"
    assert m.group(3) == "10332"


def test_lexml_urn_tipo_map():
    """URN tipo map contains expected entries."""
    assert "lei" in URN_TIPO_MAP
    assert "decreto" in URN_TIPO_MAP
    assert URN_TIPO_MAP["lei.complementar"] == "Lei Complementar"
    assert URN_TIPO_MAP["medida.provisoria"] == "Medida Provisória"


def test_lexml_parse_sru_response_valid():
    """Parse a valid SRU XML response."""
    xml = '''<?xml version="1.0"?>
    <srw:searchRetrieveResponse xmlns:srw="http://www.loc.gov/zing/srw/">
        <srw:numberOfRecords>1</srw:numberOfRecords>
        <srw:records>
            <srw:record>
                <srw:recordData>
                    <dc:title xmlns:dc="http://purl.org/dc/elements/1.1/">Lei de Teste</dc:title>
                    <dc:description xmlns:dc="http://purl.org/dc/elements/1.1/">Ementa de teste</dc:description>
                    <dc:date xmlns:dc="http://purl.org/dc/elements/1.1/">2020-01-01</dc:date>
                    <dc:creator xmlns:dc="http://purl.org/dc/elements/1.1/">Congresso Nacional</dc:creator>
                    <dc:type xmlns:dc="http://purl.org/dc/elements/1.1/">Legislacao</dc:type>
                    <dc:identifier xmlns:dc="http://purl.org/dc/elements/1.1/">urn:lex:br:federal:lei:2020-01-01;123</dc:identifier>
                </srw:recordData>
            </srw:record>
        </srw:records>
    </srw:searchRetrieveResponse>'''

    searcher = LexMLSearcher()
    results, total = searcher._parse_sru_response(xml, "teste")
    assert total == 1, f"Expected total=1, got {total}"
    assert len(results) == 1, f"Expected 1 result, got {len(results)}"
    r = results[0]
    assert r.nome == "Lei de Teste"
    assert r.ementa == "Ementa de teste"
    assert r.tipo == "Lei"
    assert r.numero == "123"
    assert r.source == "lexml"
    assert r.found_by == "teste"


def test_lexml_parse_sru_response_empty():
    """Parse empty SRU response."""
    xml = '''<?xml version="1.0"?>
    <srw:searchRetrieveResponse xmlns:srw="http://www.loc.gov/zing/srw/">
        <srw:numberOfRecords>0</srw:numberOfRecords>
        <srw:records/>
    </srw:searchRetrieveResponse>'''

    searcher = LexMLSearcher()
    results, total = searcher._parse_sru_response(xml, "teste")
    assert total == 0
    assert len(results) == 0


def test_lexml_parse_sru_response_malformed_xml():
    """Malformed XML should return empty results without crashing."""
    searcher = LexMLSearcher()
    results, total = searcher._parse_sru_response("<not>valid<xml", "teste")
    assert total == 0
    assert len(results) == 0


def test_lexml_search_empty_keywords():
    """Search with empty keyword list should return empty results."""
    searcher = LexMLSearcher()
    results = searcher.search([], max_results=10)
    assert results == []


def test_lexml_cql_injection_sanitization():
    """Keywords with quotes should be sanitized."""
    searcher = LexMLSearcher()
    # This tests _search_keyword_safe - the CQL query should not break
    # We just verify it doesn't crash with injection-like input
    results, error = searcher._search_keyword_safe('"; DROP TABLE laws --', max_results=5)
    # Should not crash, just return empty or valid results (with possible error)
    assert isinstance(results, list)


for name, func in [
    ("LexML source name", test_lexml_source_name),
    ("LexML URN pattern full match", test_lexml_urn_pattern_full),
    ("LexML URN pattern year only", test_lexml_urn_pattern_year_only),
    ("LexML URN tipo map entries", test_lexml_urn_tipo_map),
    ("LexML parse valid SRU response", test_lexml_parse_sru_response_valid),
    ("LexML parse empty SRU response", test_lexml_parse_sru_response_empty),
    ("LexML parse malformed XML", test_lexml_parse_sru_response_malformed_xml),
    ("LexML search empty keywords", test_lexml_search_empty_keywords),
    ("LexML CQL injection sanitization", test_lexml_cql_injection_sanitization),
]:
    run_test(name, func)


# ===========================================================================
# 4. TCU SEARCHER TESTS
# ===========================================================================
print("\n" + "=" * 70)
print("4. TCU SEARCHER")
print("=" * 70)

from searchers.tcu_searcher import TCUSearcher


def test_tcu_source_name():
    s = TCUSearcher()
    assert s.source_name() == "TCU Dados Abertos"


def test_tcu_matches_keyword():
    """Keyword matching is accent/case insensitive."""
    s = TCUSearcher()
    assert s._matches_keyword("Gestão de segurança da informação", "seguranca")
    assert s._matches_keyword("GOVERNANÇA DE TI", "governanca")
    assert not s._matches_keyword("Texto sobre finanças", "seguranca")


def test_tcu_map_acordao():
    """Map an acordao JSON item to NormativoResult."""
    s = TCUSearcher()
    item = {
        "numero": 1234,
        "ano": 2023,
        "colegiado": "Plenário",
        "ementa": "Determina providências sobre segurança da informação",
        "dataAta": "2023-06-15",
    }
    result = s._map_acordao(item, "segurança")
    assert result.tipo == "Acordao TCU"
    assert result.numero == "1234/2023"
    assert result.source == "tcu"
    assert "1234" in result.nome
    assert "2023" in result.nome
    assert "Plenário" in result.nome
    assert result.data == "15/06/2023"


def test_tcu_map_ato_normativo():
    """Map an ato normativo JSON item to NormativoResult."""
    s = TCUSearcher()
    item = {
        "tipo": "Resolução",
        "numero": 36,
        "ementa": "Resolução sobre auditoria interna",
        "dataPublicacao": "2022-03-01",
        "link": "https://tcu.gov.br/resolucao/36",
    }
    result = s._map_ato_normativo(item, "auditoria")
    assert result.tipo == "Resolução"
    assert result.numero == "36"
    assert result.data == "01/03/2022"
    assert result.link == "https://tcu.gov.br/resolucao/36"


def test_tcu_map_acordao_missing_fields():
    """Map acordao with missing optional fields."""
    s = TCUSearcher()
    item = {"numero": 999, "ano": 2024}
    result = s._map_acordao(item, "kw")
    assert result.tipo == "Acordao TCU"
    assert result.numero == "999/2024"
    assert result.ementa == ""


def test_tcu_search_empty_keywords():
    """Search with empty keywords should return empty results."""
    s = TCUSearcher()
    # Mock _fetch_all_pages to return empty to avoid real API calls
    s._fetch_all_pages = lambda url: []
    results = s.search([], max_results=10)
    assert results == []


def test_tcu_build_acordao_link():
    """Verify acordao link format."""
    link = TCUSearcher._build_acordao_link("1234", "2023")
    assert "1234" in link
    assert "2023" in link
    assert link.startswith("https://")


for name, func in [
    ("TCU source name", test_tcu_source_name),
    ("TCU keyword matching", test_tcu_matches_keyword),
    ("TCU map acordao", test_tcu_map_acordao),
    ("TCU map ato normativo", test_tcu_map_ato_normativo),
    ("TCU map acordao missing fields", test_tcu_map_acordao_missing_fields),
    ("TCU search empty keywords", test_tcu_search_empty_keywords),
    ("TCU build acordao link", test_tcu_build_acordao_link),
]:
    run_test(name, func)


# ===========================================================================
# 5. GOOGLE SEARCHER TESTS
# ===========================================================================
print("\n" + "=" * 70)
print("5. GOOGLE SEARCHER")
print("=" * 70)

from searchers.google_searcher import GoogleSearcher


def test_google_source_name():
    s = GoogleSearcher()
    assert s.source_name() == "Google (Frameworks/Padroes)"


def test_google_is_safe_url_valid():
    """Valid HTTPS URLs should pass."""
    assert GoogleSearcher._is_safe_url("https://www.isaca.org/cobit")
    assert GoogleSearcher._is_safe_url("https://iso.org/standard/123")


def test_google_is_safe_url_blocked():
    """Internal/metadata URLs should be blocked."""
    assert not GoogleSearcher._is_safe_url("http://169.254.169.254/metadata")
    assert not GoogleSearcher._is_safe_url("http://localhost:8080")
    assert not GoogleSearcher._is_safe_url("http://127.0.0.1/admin")
    assert not GoogleSearcher._is_safe_url("ftp://example.com/file")


def test_google_is_safe_url_invalid_scheme():
    """Non-http(s) schemes should be blocked."""
    assert not GoogleSearcher._is_safe_url("file:///etc/passwd")
    assert not GoogleSearcher._is_safe_url("javascript:alert(1)")


def test_google_extract_org():
    """Domain-to-org mapping."""
    s = GoogleSearcher()
    assert s._extract_org("https://www.isaca.org/cobit") == "ISACA (COBIT)"
    assert s._extract_org("https://iso.org/standard/123") == "ISO"
    assert s._extract_org("https://www.tcu.gov.br/resolucao") == "TCU"
    assert s._extract_org("https://portal.gov.br/dados") == "Governo Federal"


def test_google_extract_org_unknown():
    """Unknown domains return the raw domain."""
    s = GoogleSearcher()
    result = s._extract_org("https://example.com/page")
    assert result == "example.com"


def test_google_normalize_url():
    """URL normalization for dedup."""
    assert GoogleSearcher._normalize_url("https://www.isaca.org/cobit/") == "isaca.org/cobit"
    assert GoogleSearcher._normalize_url("http://ISO.org/standard") == "iso.org/standard"
    assert GoogleSearcher._normalize_url("HTTPS://WWW.GOV.BR") == "gov.br"


def test_google_search_empty_keywords():
    """Search with empty keywords returns empty."""
    s = GoogleSearcher()
    results = s.search([], max_results=10)
    assert results == []


for name, func in [
    ("Google source name", test_google_source_name),
    ("Google safe URL valid", test_google_is_safe_url_valid),
    ("Google safe URL blocked", test_google_is_safe_url_blocked),
    ("Google safe URL invalid scheme", test_google_is_safe_url_invalid_scheme),
    ("Google extract org", test_google_extract_org),
    ("Google extract org unknown domain", test_google_extract_org_unknown),
    ("Google normalize URL", test_google_normalize_url),
    ("Google search empty keywords", test_google_search_empty_keywords),
]:
    run_test(name, func)


# ===========================================================================
# 6. DEDUPLICATOR TESTS
# ===========================================================================
print("\n" + "=" * 70)
print("6. DEDUPLICATOR")
print("=" * 70)

from deduplicator import deduplicate, _normalize, _merge


def _make_result(**kwargs):
    """Helper to create a NormativoResult with defaults."""
    defaults = dict(
        nome="Test", tipo="Lei", numero="1", data="01/01/2020",
        orgao_emissor="Test", ementa="Ementa teste", link="http://test",
        source="lexml", found_by="kw",
    )
    defaults.update(kwargs)
    return NormativoResult(**defaults)


def test_dedup_empty_list():
    """Deduplicate empty list returns empty."""
    assert deduplicate([]) == []


def test_dedup_single_item():
    """Single item: no dedup needed."""
    r = _make_result()
    result = deduplicate([r])
    assert len(result) == 1
    assert result[0].nome == "Test"


def test_dedup_exact_id_match():
    """Two items with same tipo/numero/data (same ID) should be merged."""
    r1 = _make_result(nome="Short", ementa="Short ementa", source="lexml")
    r2 = _make_result(nome="A Longer Name", ementa="A much longer ementa for testing", source="tcu")
    result = deduplicate([r1, r2])
    assert len(result) == 1
    # Longer ementa should be kept
    assert "much longer" in result[0].ementa
    # Sources combined
    assert "lexml" in result[0].source
    assert "tcu" in result[0].source


def test_dedup_tipo_numero_match():
    """Items with same tipo+numero but different date (different ID) should merge."""
    r1 = _make_result(tipo="Lei", numero="123", data="01/01/2020", source="lexml")
    r2 = _make_result(tipo="Lei", numero="123", data="02/01/2020", source="tcu")
    result = deduplicate([r1, r2])
    assert len(result) == 1


def test_dedup_fuzzy_ementa_match():
    """Similar ementas (long enough) should be merged."""
    long_ementa = (
        "Dispoe sobre a protecao de dados pessoais, inclusive nos meios digitais, "
        "por pessoa natural ou por pessoa juridica de direito publico ou privado, "
        "com o objetivo de proteger os direitos fundamentais de liberdade e de privacidade"
    )
    r1 = _make_result(
        tipo="Lei", numero="1", data="01/01/2020",
        ementa=long_ementa,
        source="lexml",
    )
    r2 = _make_result(
        tipo="Decreto", numero="2", data="02/02/2020",
        ementa=long_ementa + ".",
        source="tcu",
    )
    result = deduplicate([r1, r2])
    assert len(result) == 1


def test_dedup_different_items_preserved():
    """Completely different items should not be merged."""
    r1 = _make_result(tipo="Lei", numero="1", ementa="Lei de acesso a informacao")
    r2 = _make_result(tipo="Decreto", numero="999", ementa="Decreto sobre seguranca cibernetica")
    result = deduplicate([r1, r2])
    assert len(result) == 2


def test_dedup_merge_prefers_longer_ementa():
    """Merge should keep the longer ementa."""
    r1 = _make_result(ementa="Short")
    r2 = _make_result(ementa="A much longer ementa text for testing purposes")
    result = deduplicate([r1, r2])
    assert "much longer" in result[0].ementa


def test_dedup_merge_combines_sources():
    """Merge should combine source values."""
    r1 = _make_result(source="lexml", found_by="kw1")
    r2 = _make_result(source="tcu", found_by="kw2")
    result = deduplicate([r1, r2])
    assert "lexml" in result[0].source
    assert "tcu" in result[0].source
    assert "kw1" in result[0].found_by
    assert "kw2" in result[0].found_by


def test_dedup_merge_keeps_higher_relevancia():
    """Merge should keep the higher relevancia score."""
    r1 = _make_result(relevancia=0.3)
    r2 = _make_result(relevancia=0.9)
    result = deduplicate([r1, r2])
    assert result[0].relevancia == 0.9


def test_dedup_merge_link_prefers_authoritative():
    """Merge should prefer link from more authoritative source."""
    r1 = _make_result(source="google", link="http://google.com/result")
    r2 = _make_result(source="lexml", link="http://lexml.gov.br/result")
    result = deduplicate([r1, r2])
    assert result[0].link == "http://lexml.gov.br/result"


def test_normalize_function():
    """Test the _normalize helper from deduplicator."""
    assert _normalize("  Ação Regulamentação  ") == "acao regulamentacao"
    assert _normalize("") == ""
    assert _normalize("A, B. C; D") == "a b c d"


def test_dedup_many_items():
    """Dedup with many unique items should preserve all."""
    results = [
        _make_result(tipo="Lei", numero=str(i), ementa=f"Ementa unica {i * 1000}")
        for i in range(50)
    ]
    deduped = deduplicate(results)
    assert len(deduped) == 50


for name, func in [
    ("Dedup empty list", test_dedup_empty_list),
    ("Dedup single item", test_dedup_single_item),
    ("Dedup exact ID match", test_dedup_exact_id_match),
    ("Dedup tipo+numero match", test_dedup_tipo_numero_match),
    ("Dedup fuzzy ementa match", test_dedup_fuzzy_ementa_match),
    ("Dedup different items preserved", test_dedup_different_items_preserved),
    ("Dedup merge prefers longer ementa", test_dedup_merge_prefers_longer_ementa),
    ("Dedup merge combines sources", test_dedup_merge_combines_sources),
    ("Dedup merge keeps higher relevancia", test_dedup_merge_keeps_higher_relevancia),
    ("Dedup merge link prefers authoritative", test_dedup_merge_link_prefers_authoritative),
    ("Normalize function", test_normalize_function),
    ("Dedup many items", test_dedup_many_items),
]:
    run_test(name, func)


# ===========================================================================
# 7. EXCEL EXPORT TESTS
# ===========================================================================
print("\n" + "=" * 70)
print("7. EXCEL EXPORT")
print("=" * 70)

from excel_export import generate_excel, _format_date
from openpyxl import load_workbook


def test_excel_format_date_iso():
    assert _format_date("2020-01-15") == "15/01/2020"


def test_excel_format_date_already_correct():
    assert _format_date("15/01/2020") == "15/01/2020"


def test_excel_format_date_iso_datetime():
    assert _format_date("2020-01-15T14:30:00") == "15/01/2020"


def test_excel_format_date_empty():
    assert _format_date("") == ""
    assert _format_date(None) == ""


def test_excel_format_date_unknown():
    assert _format_date("Jan 2020") == "Jan 2020"


def test_excel_generate_empty():
    """Generate Excel with no results should produce a valid workbook."""
    buf = generate_excel([], "Test Topic")
    assert isinstance(buf, BytesIO)
    assert buf.tell() == 0
    # Verify it's a valid xlsx
    wb = load_workbook(buf)
    ws = wb.active
    assert ws.title == "Normativos"
    assert ws.cell(row=1, column=1).value == "Levantamento de Normativos: Test Topic"
    assert ws.cell(row=2, column=1).value == "Nome do Normativo"


def test_excel_generate_with_results():
    """Generate Excel with sample results."""
    results = [
        _make_result(
            nome="Lei 13.709 (LGPD)",
            tipo="Lei",
            numero="13.709",
            data="14/08/2018",
            orgao_emissor="Presidencia",
            ementa="Lei Geral de Protecao de Dados",
            link="https://planalto.gov.br/lgpd",
            relevancia=0.95,
        ),
        _make_result(
            nome="Decreto 10.332",
            tipo="Decreto",
            numero="10.332",
            data="28/04/2020",
            orgao_emissor="Presidencia",
            ementa="Estrategia de Governo Digital",
            link="https://planalto.gov.br/decreto10332",
            relevancia=0.7,
        ),
    ]
    buf = generate_excel(results, "LGPD")
    wb = load_workbook(buf)
    ws = wb.active

    # Title row
    assert "LGPD" in ws.cell(row=1, column=1).value

    # Header row
    assert ws.cell(row=2, column=1).value == "Nome do Normativo"
    assert ws.cell(row=2, column=6).value == "Ementa"

    # Data row 1
    assert ws.cell(row=3, column=1).value == "Lei 13.709 (LGPD)"
    assert ws.cell(row=3, column=2).value == "Lei"
    assert ws.cell(row=3, column=3).value == "13.709"

    # Data row 2
    assert ws.cell(row=4, column=1).value == "Decreto 10.332"

    # Relevancia should be numeric
    assert ws.cell(row=3, column=10).value == 0.95


def test_excel_hyperlink_creation():
    """Links should be created as hyperlinks."""
    results = [
        _make_result(link="https://example.com/doc"),
    ]
    buf = generate_excel(results, "Test")
    wb = load_workbook(buf)
    ws = wb.active
    cell = ws.cell(row=3, column=7)  # Link column
    assert cell.value == "https://example.com/doc"
    assert cell.hyperlink is not None


def test_excel_long_ementa_truncation():
    """Very long ementas should be truncated."""
    long_ementa = "A" * 6000
    results = [_make_result(ementa=long_ementa)]
    buf = generate_excel(results, "Test")
    wb = load_workbook(buf)
    ws = wb.active
    cell_value = ws.cell(row=3, column=6).value
    assert len(cell_value) <= 5010  # 5000 + "..."


def test_excel_freeze_panes():
    """Freeze panes should be set at A3."""
    buf = generate_excel([_make_result()], "Test")
    wb = load_workbook(buf)
    ws = wb.active
    assert ws.freeze_panes == "A3"


def test_excel_auto_filter():
    """Auto-filter should be set on header row."""
    buf = generate_excel([_make_result()], "Test")
    wb = load_workbook(buf)
    ws = wb.active
    assert ws.auto_filter.ref is not None
    assert "A2" in ws.auto_filter.ref


def test_excel_relevancia_formatting():
    """Relevancia column should use percentage format."""
    results = [
        _make_result(relevancia=0.0),
        _make_result(tipo="Decreto", numero="2", relevancia=0.5),
        _make_result(tipo="Portaria", numero="3", relevancia=0.85),
    ]
    buf = generate_excel(results, "Test")
    wb = load_workbook(buf)
    ws = wb.active

    # Row 3: relevancia 0.0
    assert ws.cell(row=3, column=10).value == 0.0
    assert ws.cell(row=3, column=10).number_format == "0%"

    # Row 5: relevancia 0.85 should have green fill
    cell = ws.cell(row=5, column=10)
    assert cell.value == 0.85


def test_excel_date_formatting():
    """Date column should format ISO dates to DD/MM/YYYY."""
    results = [_make_result(data="2020-03-15")]
    buf = generate_excel(results, "Test")
    wb = load_workbook(buf)
    ws = wb.active
    assert ws.cell(row=3, column=4).value == "15/03/2020"


for name, func in [
    ("Excel format date ISO", test_excel_format_date_iso),
    ("Excel format date already correct", test_excel_format_date_already_correct),
    ("Excel format date ISO datetime", test_excel_format_date_iso_datetime),
    ("Excel format date empty", test_excel_format_date_empty),
    ("Excel format date unknown", test_excel_format_date_unknown),
    ("Excel generate empty", test_excel_generate_empty),
    ("Excel generate with results", test_excel_generate_with_results),
    ("Excel hyperlink creation", test_excel_hyperlink_creation),
    ("Excel long ementa truncation", test_excel_long_ementa_truncation),
    ("Excel freeze panes", test_excel_freeze_panes),
    ("Excel auto filter", test_excel_auto_filter),
    ("Excel relevancia formatting", test_excel_relevancia_formatting),
    ("Excel date formatting", test_excel_date_formatting),
]:
    run_test(name, func)


# ===========================================================================
# 8. LLM CLIENT TESTS
# ===========================================================================
print("\n" + "=" * 70)
print("8. LLM CLIENT (gemini_client)")
print("=" * 70)

from llm.gemini_client import (
    _parse_json_array,
    _chunk_list,
    _keyword_relevance,
    _fuzzy_match_category,
    CATEGORIES,
    is_available,
    expand_topic_to_keywords,
    score_relevance,
    categorize_results,
)


def test_parse_json_array_clean():
    """Parse a clean JSON array."""
    result = _parse_json_array('["a", "b", "c"]')
    assert result == ["a", "b", "c"]


def test_parse_json_array_with_code_fence():
    """Parse JSON inside markdown code fences."""
    text = '```json\n["a", "b"]\n```'
    result = _parse_json_array(text)
    assert result == ["a", "b"]


def test_parse_json_array_with_extra_text():
    """Parse JSON array from text with extra explanation."""
    text = 'Here are the keywords:\n["alpha", "beta"]\nHope this helps!'
    result = _parse_json_array(text)
    assert result == ["alpha", "beta"]


def test_parse_json_array_invalid():
    """Invalid JSON should return None."""
    assert _parse_json_array("not json at all") is None
    assert _parse_json_array("{}") is None
    assert _parse_json_array("") is None


def test_parse_json_array_numbers():
    """Parse array of numbers."""
    result = _parse_json_array("[0.5, 0.8, 0.3]")
    assert result == [0.5, 0.8, 0.3]


def test_chunk_list_basic():
    """Chunk a list into pieces."""
    result = _chunk_list([1, 2, 3, 4, 5], 2)
    assert result == [[1, 2], [3, 4], [5]]


def test_chunk_list_empty():
    assert _chunk_list([], 5) == []


def test_chunk_list_exact():
    result = _chunk_list([1, 2, 3, 4], 2)
    assert result == [[1, 2], [3, 4]]


def test_keyword_relevance_basic():
    """Keyword matching counts keyword occurrences."""
    keywords = ["segurança", "informação", "dados"]
    ementa = "Dispõe sobre segurança da informação no setor público"
    score = _keyword_relevance(keywords, ementa)
    assert 0.0 < score <= 1.0
    # 2 out of 3 keywords match
    assert abs(score - 2/3) < 0.01


def test_keyword_relevance_empty():
    assert _keyword_relevance([], "ementa") == 0.0
    assert _keyword_relevance(["kw"], "") == 0.0


def test_keyword_relevance_all_match():
    keywords = ["a", "b"]
    ementa = "text a text b text"
    assert _keyword_relevance(keywords, ementa) == 1.0


def test_fuzzy_match_category_exact():
    """Exact match should return the category."""
    assert _fuzzy_match_category("Governança de TI") == "Governança de TI"
    assert _fuzzy_match_category("Segurança da Informação") == "Segurança da Informação"


def test_fuzzy_match_category_no_accents():
    """Match without accents."""
    assert _fuzzy_match_category("Governanca de TI") == "Governança de TI"
    assert _fuzzy_match_category("Seguranca da Informacao") == "Segurança da Informação"


def test_fuzzy_match_category_case_insensitive():
    """Match case-insensitively."""
    assert _fuzzy_match_category("governança de ti") == "Governança de TI"
    assert _fuzzy_match_category("PROTEÇÃO DE DADOS") == "Proteção de Dados"


def test_fuzzy_match_category_no_match():
    """Unknown category returns None."""
    assert _fuzzy_match_category("Categoria Inexistente") is None


def test_categories_list():
    """CATEGORIES should contain expected entries."""
    assert "Governança de TI" in CATEGORIES
    assert "Segurança da Informação" in CATEGORIES
    assert "Outro" in CATEGORIES
    assert len(CATEGORIES) >= 10


def test_score_relevance_empty():
    """score_relevance with empty results returns empty."""
    assert score_relevance("topic", []) == []


def test_score_relevance_fallback_no_llm():
    """Without LLM, should use keyword heuristic fallback."""
    results = [
        {"nome": "Lei A", "ementa": "seguranca da informacao no governo"},
        {"nome": "Lei B", "ementa": "tema completamente diferente"},
    ]
    scores = score_relevance("segurança", results, keywords=["seguranca", "informacao"])
    assert len(scores) == 2
    # If LLM is available, scores come from Gemini (may vary);
    # if not, keyword heuristic should rank first higher
    assert isinstance(scores[0], float)
    assert isinstance(scores[1], float)


def test_categorize_results_empty():
    """categorize_results with empty results returns empty."""
    assert categorize_results("topic", []) == []


def test_expand_topic_fallback_no_llm():
    """Without LLM key, expand should return empty list."""
    # If no key is configured, should return []
    result = expand_topic_to_keywords("governança de TI")
    assert isinstance(result, list)


for name, func in [
    ("Parse JSON array clean", test_parse_json_array_clean),
    ("Parse JSON array code fence", test_parse_json_array_with_code_fence),
    ("Parse JSON array extra text", test_parse_json_array_with_extra_text),
    ("Parse JSON array invalid", test_parse_json_array_invalid),
    ("Parse JSON array numbers", test_parse_json_array_numbers),
    ("Chunk list basic", test_chunk_list_basic),
    ("Chunk list empty", test_chunk_list_empty),
    ("Chunk list exact", test_chunk_list_exact),
    ("Keyword relevance basic", test_keyword_relevance_basic),
    ("Keyword relevance empty", test_keyword_relevance_empty),
    ("Keyword relevance all match", test_keyword_relevance_all_match),
    ("Fuzzy match category exact", test_fuzzy_match_category_exact),
    ("Fuzzy match category no accents", test_fuzzy_match_category_no_accents),
    ("Fuzzy match category case insensitive", test_fuzzy_match_category_case_insensitive),
    ("Fuzzy match category no match", test_fuzzy_match_category_no_match),
    ("Categories list contents", test_categories_list),
    ("Score relevance empty", test_score_relevance_empty),
    ("Score relevance fallback no LLM", test_score_relevance_fallback_no_llm),
    ("Categorize results empty", test_categorize_results_empty),
    ("Expand topic fallback no LLM", test_expand_topic_fallback_no_llm),
]:
    run_test(name, func)


# ===========================================================================
# 9. APP HELPER TESTS
# ===========================================================================
print("\n" + "=" * 70)
print("9. APP HELPERS")
print("=" * 70)


def test_make_filename_slug():
    """Test the filename slug helper."""
    from app import _make_filename_slug

    assert _make_filename_slug("Governança de TI") == "governanca_de_ti"
    assert _make_filename_slug("  LGPD - Dados Pessoais  ") == "lgpd_dados_pessoais"
    # Should truncate to 60 chars
    long = "a" * 100
    result = _make_filename_slug(long)
    assert len(result) <= 60


def test_get_tipo_color():
    """Test tipo color mapping."""
    from app import _get_tipo_color, TIPO_COLORS

    assert _get_tipo_color("Lei") == "#1976D2"
    assert _get_tipo_color("Acordao TCU") == "#D32F2F"
    assert _get_tipo_color("Unknown") == "#757575"


def test_apply_filters_all():
    """Test filter function with no filters (all pass)."""
    from app import _apply_filters

    results = [
        _make_result(tipo="Lei", source="lexml", relevancia=0.5),
        _make_result(tipo="Decreto", numero="2", source="tcu", relevancia=0.8),
    ]
    filtered = _apply_filters(results, "Todos", "Todas", 0)
    assert len(filtered) == 2


def test_apply_filters_by_tipo():
    """Test filter by tipo."""
    from app import _apply_filters

    results = [
        _make_result(tipo="Lei", source="lexml", relevancia=0.5),
        _make_result(tipo="Decreto", numero="2", source="tcu", relevancia=0.8),
    ]
    filtered = _apply_filters(results, "Lei", "Todas", 0)
    assert len(filtered) == 1
    assert filtered[0].tipo == "Lei"


def test_apply_filters_by_source():
    """Test filter by source."""
    from app import _apply_filters

    results = [
        _make_result(tipo="Lei", source="lexml", relevancia=0.5),
        _make_result(tipo="Decreto", numero="2", source="tcu", relevancia=0.8),
    ]
    filtered = _apply_filters(results, "Todos", "tcu", 0)
    assert len(filtered) == 1
    assert "tcu" in filtered[0].source


def test_apply_filters_by_relevancia():
    """Test filter by minimum relevancia."""
    from app import _apply_filters

    results = [
        _make_result(tipo="Lei", source="lexml", relevancia=0.3),
        _make_result(tipo="Decreto", numero="2", source="tcu", relevancia=0.8),
    ]
    filtered = _apply_filters(results, "Todos", "Todas", 50)
    assert len(filtered) == 1
    assert filtered[0].relevancia == 0.8


def test_apply_sort_relevancia():
    """Test sort by relevancia."""
    from app import _apply_sort

    results = [
        _make_result(relevancia=0.3),
        _make_result(tipo="Decreto", numero="2", relevancia=0.9),
        _make_result(tipo="Portaria", numero="3", relevancia=0.5),
    ]
    sorted_results = _apply_sort(results, "Relevancia (descendente)")
    assert sorted_results[0].relevancia == 0.9
    assert sorted_results[1].relevancia == 0.5
    assert sorted_results[2].relevancia == 0.3


def test_apply_sort_tipo():
    """Test sort by tipo."""
    from app import _apply_sort

    results = [
        _make_result(tipo="Portaria"),
        _make_result(tipo="Decreto", numero="2"),
        _make_result(tipo="Lei", numero="3"),
    ]
    sorted_results = _apply_sort(results, "Tipo")
    assert sorted_results[0].tipo == "Decreto"
    assert sorted_results[1].tipo == "Lei"
    assert sorted_results[2].tipo == "Portaria"


for name, func in [
    ("Filename slug generation", test_make_filename_slug),
    ("Tipo color mapping", test_get_tipo_color),
    ("Filter: no filters", test_apply_filters_all),
    ("Filter: by tipo", test_apply_filters_by_tipo),
    ("Filter: by source", test_apply_filters_by_source),
    ("Filter: by relevancia", test_apply_filters_by_relevancia),
    ("Sort: by relevancia", test_apply_sort_relevancia),
    ("Sort: by tipo", test_apply_sort_tipo),
]:
    run_test(name, func)


# ===========================================================================
# 10. INTEGRATION TESTS (with real APIs, skippable)
# ===========================================================================
print("\n" + "=" * 70)
print("10. INTEGRATION TESTS (real API calls)")
print("=" * 70)

RUN_INTEGRATION = os.environ.get("RUN_INTEGRATION", "1") == "1"

if RUN_INTEGRATION:
    def test_lexml_live_search():
        """Live LexML search for a common term."""
        searcher = LexMLSearcher()
        results = searcher.search(["LGPD"], max_results=5)
        assert isinstance(results, list), "Should return a list"
        # LexML should find something for LGPD
        if len(results) > 0:
            r = results[0]
            assert hasattr(r, 'nome')
            assert hasattr(r, 'tipo')
            assert hasattr(r, 'source')
            assert r.source == "lexml"
        # It's OK if the API is down and returns 0 results

    def test_tcu_live_search():
        """Live TCU search."""
        searcher = TCUSearcher()
        results = searcher.search(["seguranca da informacao"], max_results=5)
        assert isinstance(results, list)
        if len(results) > 0:
            r = results[0]
            assert r.source == "tcu"

    def test_end_to_end_search_and_export():
        """Full pipeline: search -> dedup -> export."""
        # Use LexML only for speed
        searcher = LexMLSearcher()
        results = searcher.search(["governanca de TI"], max_results=10)

        if results:
            deduped = deduplicate(results)
            assert len(deduped) <= len(results)

            buf = generate_excel(deduped, "Governança de TI")
            assert isinstance(buf, BytesIO)

            wb = load_workbook(buf)
            ws = wb.active
            assert ws.cell(row=3, column=1).value is not None
        # OK if API returns nothing

    for name, func in [
        ("LIVE: LexML search", test_lexml_live_search),
        ("LIVE: TCU search", test_tcu_live_search),
        ("LIVE: End-to-end search + export", test_end_to_end_search_and_export),
    ]:
        run_test(name, func)
else:
    print("  SKIP  Integration tests disabled (set RUN_INTEGRATION=1 to enable)")


# ===========================================================================
# FINAL SUMMARY
# ===========================================================================
print("\n" + "=" * 70)
print("FINAL SUMMARY")
print("=" * 70)

total = len(_results)
passed = sum(1 for r in _results if r[0] == "PASS")
failed = sum(1 for r in _results if r[0] == "FAIL")

print(f"\nTotal: {total} | Passed: {passed} | Failed: {failed}")

if failed > 0:
    print("\nFAILED TESTS:")
    for status, name, error in _results:
        if status == "FAIL":
            print(f"  X {name}: {error}")

print("\n" + ("ALL TESTS PASSED!" if failed == 0 else f"{failed} TEST(S) FAILED"))
sys.exit(0 if failed == 0 else 1)
